import logging
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)


def export_excel(jobs: list, contacts: dict, output_dir: str, run_stats: dict = None) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return ""

    Path(output_dir).mkdir(parents=True, exist_ok=True)
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    out_path = Path(output_dir) / f"jobs_{date_str}.xlsx"

    wb = openpyxl.Workbook()

    # ── Sheet 1: Jobs ─────────────────────────────────────────────────
    ws_jobs = wb.active
    ws_jobs.title = "Jobs"

    job_headers = [
        "Title", "Company", "Location", "Salary", "Salary Min", "Salary Max",
        "Salary Period", "Job Type", "Posted", "Expires", "Apply URL", "Source(s)", "Scraped At"
    ]
    _write_header_row(ws_jobs, job_headers)

    for job in jobs:
        sources = "|".join(job.sources or [job.source])
        row = [
            job.title,
            job.company,
            job.location,
            job.salary_text,
            job.salary_min,
            job.salary_max,
            job.salary_period,
            job.job_type,
            job.posted_at,
            job.expires_at,
            job.apply_url,
            sources,
            job.scraped_at,
        ]
        ws_jobs.append(row)

    # Hyperlink apply URLs
    apply_col_idx = job_headers.index("Apply URL") + 1
    for row_idx, job in enumerate(jobs, start=2):
        if job.apply_url:
            cell = ws_jobs.cell(row=row_idx, column=apply_col_idx)
            cell.hyperlink = job.apply_url
            cell.font = Font(color="0563C1", underline="single")

    _auto_width(ws_jobs)
    ws_jobs.freeze_panes = "A2"

    # ── Sheet 2: Contacts ─────────────────────────────────────────────
    ws_contacts = wb.create_sheet("Contacts")

    contact_headers = [
        "Company", "Phone(s)", "Email(s)", "Contact Person", "Address",
        "Website", "Company Number", "Company Type", "Confidence Score",
        "AI Used", "Enrichment Sources"
    ]
    _write_header_row(ws_contacts, contact_headers)

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    amber_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    for row_idx, (company, contact) in enumerate(contacts.items(), start=2):
        confidence = contact.confidence_score
        row = [
            contact.company,
            " | ".join(contact.phone_numbers),
            " | ".join(contact.emails),
            contact.contact_person,
            contact.address,
            contact.company_website,
            contact.company_number,
            contact.company_type,
            confidence,
            "Yes" if contact.ai_used else "No",
            " | ".join(contact.enrichment_sources),
        ]
        ws_contacts.append(row)

        # Colour AI Used column (index 10, 1-based col 10+1=11... but header index 9)
        ai_col = contact_headers.index("AI Used") + 1
        if contact.ai_used:
            ws_contacts.cell(row=row_idx, column=ai_col).fill = yellow_fill

        # Colour confidence score
        conf_col = contact_headers.index("Confidence Score") + 1
        conf_cell = ws_contacts.cell(row=row_idx, column=conf_col)
        if confidence >= 70:
            conf_cell.fill = green_fill
        elif confidence >= 40:
            conf_cell.fill = amber_fill
        else:
            conf_cell.fill = red_fill

        # Hyperlink website
        web_col = contact_headers.index("Website") + 1
        if contact.company_website:
            cell = ws_contacts.cell(row=row_idx, column=web_col)
            cell.hyperlink = contact.company_website
            cell.font = Font(color="0563C1", underline="single")

    _auto_width(ws_contacts)
    ws_contacts.freeze_panes = "A2"

    # ── Sheet 3: Summary ─────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    _write_header_row(ws_summary, ["Metric", "Value"])

    stats = run_stats or {}
    ai_calls = stats.get("ai_calls", 0)

    by_source = {}
    for job in jobs:
        src = job.source
        by_source[src] = by_source.get(src, 0) + 1

    jobs_with_phone = sum(1 for j in jobs if j.company and contacts.get(j.company) and contacts[j.company].phone_numbers)
    jobs_with_email = sum(1 for j in jobs if j.company and contacts.get(j.company) and contacts[j.company].emails)
    jobs_with_contact = sum(1 for j in jobs if j.company and contacts.get(j.company) and (contacts[j.company].phone_numbers or contacts[j.company].emails))

    summary_rows = [
        ("Run Timestamp", datetime.utcnow().isoformat() + "Z"),
        ("Total Jobs Scraped", len(jobs)),
        ("Unique Companies Enriched", len(contacts)),
        ("Jobs With Any Contact Info", jobs_with_contact),
        ("Jobs With Phone Number", jobs_with_phone),
        ("Jobs With Email", jobs_with_email),
        ("AI Calls Made", ai_calls),
        ("", ""),
    ]
    for src, count in sorted(by_source.items()):
        summary_rows.append((f"Jobs from {src}", count))

    for row in summary_rows:
        ws_summary.append(row)

    _auto_width(ws_summary)
    ws_summary.freeze_panes = "A2"

    wb.save(out_path)
    logger.info(f"Excel export: {len(jobs)} jobs → {out_path}")
    return str(out_path)


def _write_header_row(ws, headers: list):
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if col_letter is None:
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(cell.column)
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
